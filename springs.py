import os
import openpyxl
from math import ceil


class Files:
    def __init__(self) -> None:
        # self.cur_dir = os.path.dirname(os.path.realpath(__file__))
        self.file_cost_name = "СебестоимостьТРЭМ-АГВ+Прайс расчет (калькулятор).xlsx"
        # self.file_cost_path = self.find(self.file_cost_name, self.cur_dir)
        self.file_cost_path = self.file_cost_name

    def find(self, name, path):
        for root, dirs, files in os.walk(path):
            if name in files:
                return os.path.join(root, name)

    def load_xl_springs(self):
        try:
            diameters = []
            wb = openpyxl.load_workbook(filename=self.file_cost_path, data_only=True)
            ws = wb["Расчет кол-ва пружин для НПЗ"]
            for row in ws.iter_rows(min_row=1, max_col=2, max_row=ws.max_row):
                if all([isinstance(i.value, int) for i in row]) and len(row) == 2:
                    diameters.append({"diameter": int(row[0].value), "num_of_springs": int(row[1].value)})
            return diameters
        except Exception as error:
            print(error)
            print("Либо файла с диаметрами нет, либо он открыт. А так я хз"
                  "У тебя лист 'Расчет кол-ва пружин для НПЗ' есть? ")
            return 0


def verification():
    while not Files().file_cost_path:
        input(
            f"Нет файла {Files().file_cost_name} в папке c программы или он назван не так. Перенесите или переименуйте и нажмине Enter: ")


def check_digits(num: str):
    for i in num:
        if not i.isdigit():
            print("Неправильный диаметр")
            return False
    if int(num) <= 0:
        print("Неправильный диаметр")
        return False
    return True


def get_num_springs(diam, diameters):
    diameters = sorted(diameters, key=lambda x: x["diameter"])
    if diam == diameters[0]["diameter"]:
        return diameters[0]["num_of_springs"]
    for diameter in diameters:
        if diam == diameter["diameter"]:
            return diameter["num_of_springs"]
        elif diam < diameter["diameter"]:
            num = ceil(diameter["num_of_springs"] / diameter["diameter"] * diam)
            return num
        elif diam > diameter["diameter"] and diameter == diameters[-1]:
            num = ceil(diameter["num_of_springs"] / diameter["diameter"] * diam)
            return num
        else:
            continue
    else:
        print("Не получилось посчитать")


def main():
    verification()
    diameters = Files().load_xl_springs()
    diam = input("Введи диаметр: ").strip()
    while True:
        if check_digits(diam):
            break
        else:
            diam = input("Введи диаметр: ").strip()
    print(f"Кол-во пружин: {get_num_springs(int(diam), diameters)} шт.")


if __name__ == "__main__":
    while True:
        main()
