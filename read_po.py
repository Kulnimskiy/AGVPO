import csv
import time

from openpyxl import Workbook, load_workbook
from datetime import datetime
from os import listdir
from os.path import isfile, join
from pprint import pprint


# gets paths for all the files in the folder
def get_files(folder_path: str):  # Путь к папке с файлами PO
    files_paths = [join(folder_path, f) for f in listdir(folder_path) if isfile(join(folder_path, f))]
    return files_paths


class PurchaseOrder:
    def __init__(self, path: str):
        self.path = path
        self._data = None
        self.__retrieve_data()
        self._order_date = None
        self.set_order_date()
        self._po_num = None
        self.set_po_num()

    @property
    def data(self):
        return self._data

    def __retrieve_data(self):
        data = []
        wb = load_workbook(filename=self.path, data_only=True, read_only=True)
        ws = wb.active
        for row in ws.iter_rows():
            row_data = []
            for cell in row:
                if isinstance(cell.value, datetime):  # convert datatime obj into the right format
                    row_data.append(cell.value.strftime("%d.%m.%Y"))
                else:
                    row_data.append(cell.value)
            data.append(row_data)
        self._data = data

    @property
    def order_date(self):
        return self._order_date

    def set_order_date(self):
        for row in self.data:
            found = False  # when we find the data
            for cell in row:
                if "дата заказа" in str(cell).lower():
                    found = True
                    continue
                if found and cell is not None:
                    self._order_date = cell
                    return

    @property
    def po_number(self):
        return self._po_num

    def set_po_num(self):
        path = self.path.split("\\")
        file_name = path[-1].strip()
        po_number = ""
        found = False
        for symbol in file_name:
            if symbol.isdigit():
                po_number += symbol
                found = True
            if symbol.isspace() and found or symbol == "(" and found:
                self._po_num = po_number
                return
        self._po_num = po_number

        # for row in self.data:
        #     found = False  # when we find the data
        #     for cell in row:
        #         if cell is not None and all([letter.isdigit() for letter in str(cell)]) and len(str(cell)) < 5:
        #             self._po_num = cell
        #             return

    def get_table(self):
        table_start_row = None
        description_index = None
        article_index = None
        drawing_index = None
        assemble_index = None
        agv_name_index = None
        quantity_index = None
        price_index = None

        for row in self.data:
            for cell in row:
                if cell is None:
                    continue
                if "описан" in str(cell).lower() and description_index is None:
                    description_index = row.index(cell)
                    table_start_row = self.data.index(row)
        if description_index is None:
            print("Столбик с описание не найден в ", self.path)
            return
        header_row = self.data[table_start_row]
        for cell in header_row:
            if "артик" in str(cell).lower() and article_index is None:
                article_index = header_row.index(cell)
            if "черте" in str(cell).lower() and drawing_index is None:
                drawing_index = header_row.index(cell)
            if "сборк" in str(cell).lower() and assemble_index is None:
                assemble_index = header_row.index(cell)
            if "наимен" in str(cell).lower() and "агв" in str(cell).lower() and agv_name_index is None:
                agv_name_index = header_row.index(cell)
            if "кол" in str(cell).lower() and quantity_index is None:
                quantity_index = header_row.index(cell)
            if "цена" in str(cell).lower() and price_index is None:
                price_index = header_row.index(cell)
            if all([drawing_index, agv_name_index, quantity_index, price_index, article_index, assemble_index]):
                break
        print()
        print(header_row)
        test_list = [drawing_index, agv_name_index, quantity_index, price_index, article_index, assemble_index]
        if not all(test_list):
            indexes = ["drawing_index", "agv_name_index", "quantity_index", "price_index", "article_index", "assemble_index"]
            print("Один или несколько столбцов не было найдено в ", self.path)
            print(list(zip(indexes, test_list)))


        table_data = []
        for row in self.data[table_start_row:]:
            row_data = {}
            if row[description_index] is not None and len(str(row[description_index])) > 10:
                row_data["description"] = row[description_index]
            else:
                continue  # if there is no description, there is no item
            row_data["article"] = row[article_index] if article_index else article_index
            row_data["drawing"] = row[drawing_index] if drawing_index else drawing_index
            row_data["assemble"] = row[assemble_index] if assemble_index else assemble_index
            row_data["agv_name"] = row[agv_name_index] if agv_name_index else agv_name_index
            row_data["quantity"] = row[quantity_index] if quantity_index else quantity_index
            row_data["price"] = row[price_index] if price_index else price_index
            row_data["date"] = self.order_date
            row_data["PO_number"] = self.po_number
            row_data["path"] = self.path
            table_data.append(row_data)

        return table_data


def save_to_csv(table_data: list[dict]):
    agv_filename = "C:\\Users\\User\\Desktop\\AGV projects\\PO_consolidator\\АГВ_все_PO.csv"
    yanos_filename = "C:\\Users\\User\\Desktop\\AGV projects\\PO_consolidator\\ЯНОС_все_PO.csv"
    with open(yanos_filename, 'w', encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(
            ("Описание", "Артикул", "Чертеж", "№_сборки", "Наименование_АГВ_конечнику", "Кол-во", "Цена_за_шт_без_НДС",
             "Дата_заказа", "Номер_PO", "Путь_к_файлу"))
        for item in table_data:
            try:
                writer.writerow((item["description"], item["article"], item["drawing"], item["assemble"],
                                 item['agv_name'], item['quantity'], item['price'],
                                 item['date'], item['PO_number'], item['path']))
            except Exception as e:
                print(e)
                print(item)

AGV_PO_PATH = "\\\\public\\Users\\Илья Земятов\\Денис Бушуев\\Документы по производству\\Заказы"
YANOS_PO_PATH = "\\\\public\\Users\\Сергей Кульминский\\Моя\\Клиенты\\Славнефть-ЯНОС\\ВСЕ PO ЯНОС"

def main():
    
    po_files = get_files(YANOS_PO_PATH)
    results = []
    for file in po_files:
        try:
            data = PurchaseOrder(file)
            results.extend(data.get_table())
        except Exception as e:
            print(e)
            print(file)
    save_to_csv(results)


def test():
    po_files = get_files(AGV_PO_PATH)

    for file in po_files:
        print(file)
        data = PurchaseOrder(file)
        pprint(data.po_number)


if __name__ == '__main__':
    main()
    print("""
THE PROCESS IS SUCCESSFUL
    """)
    input("Press ENTER to exit...")
