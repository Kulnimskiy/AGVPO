from openpyxl import load_workbook, Workbook
import xlwings
from File_search import Files


class CostFile:
    def __init__(self):
        self.path = Files().file_cost_path

    def write_diameter(self, diameter):
        """Устанавливаем диаметер для расчета в ячейке А2
        и убираем значение в ячейке В2 чтобы наличие РУ определялось
        самой программой"""
        try:
            wb = load_workbook(self.path)
            ws = wb.active
            ws["A2"].value = diameter
            ws["B2"].value = ""
            wb.save(self.path)
            wb.close()
        except Exception as error:
            print(error)

    def excel_open_close(self):
        try:
            excel_app = xlwings.App(visible=False)
            excel_book = excel_app.books.open(self.path)
            excel_book.save()
            excel_book.close()
            excel_app.quit()
        except Exception as error:
            print(error)

    def get_cost_valve(self, diameter):
        """Значения подставляем и в фоновом режиме открываем и закрываем приложение,
        чтобы после подстановки формулы экселевские сами посчиталисm"""
        self.write_diameter(diameter)
        self.excel_open_close()
        wb = load_workbook(self.path, data_only=True)
        ws = wb.active
        return ws["C2"].value

    def get_cost_zip(self, diameter):
        """Значения подставляем и в фоновом режиме открываем и закрываем приложение,
        чтобы после подстановки формулы экселевские сами посчиталисm"""
        self.write_diameter(diameter)
        self.excel_open_close()
        wb = load_workbook(self.path, data_only=True)
        ws = wb.active
        return ws["D2"].value

    def get_all_prices(self, min_diam, max_diam):
        all_prices = []
        count_down = max_diam - min_diam
        for diameter in range(min_diam, max_diam + 1):
            count_down -= 1
            print(count_down)
            self.write_diameter(diameter)
            self.excel_open_close()
            wb = load_workbook(self.path, data_only=True)
            ws = wb.active
            cost_valve, cost_zip, price_valve, price_zip = ws["C2"].value, ws["D2"].value, ws["E2"].value, ws[
                "F2"].value
            all_prices.append([diameter, cost_valve, cost_zip, price_valve, price_zip])
        return all_prices

    @staticmethod
    def write_all_prices(all_prices: list):
        """На будущее если вдруг программку надо будет ускорить или еренести все цены
        По идее, она должна их пересчитывать через какой то условленный период либо спрашивать,
         нужно ли цецы перед составлением PO посчитать"""
        wb = Workbook()
        ws = wb.active
        ws.append(["Диаметр мм", "Себе-сть Клапан", "Себе-сть ЗИП", "Прайс Клапан", "Прайс ЗИП"])
        for prices in all_prices:
            ws.append(prices)
        wb.save("C:\\Users\\sk\\Desktop\\АГВ Проекты\\PO_авто\\V3\\Все цены.xlsx")


def main():
    try:
        diameter = input("Введите расчетный диаметр: ")
        diameters = [int(diameter)]
        costs = CostFile()
        for diameter in diameters:
            print(costs.get_cost_zip(diameter))
    except Exception as error:
        print(error)


if __name__ == '__main__':
    main()
